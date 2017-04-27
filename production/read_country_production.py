from openpyxl import load_workbook
import production

wb = load_workbook(filename = 'country_production_to_import.xlsx', data_only=True)
sheet = wb.active
row_count = sheet.max_row - 1
# column_count = sheet.max_column

# production year list
year_start = 1994

#read each row for the production and save to database
row_counter = 1
cell_range = sheet['A2': 'JS' + str(row_count)]
for row in cell_range: # This is iterating through rows 1-7
    country_id = None
    country_name = None
    production_type = None
    counter = 0
    prod_counter = 0
    for cell in row: # This iterates through the columns(cells) in that row
    	if counter == 0:
    		country_id = int(cell.value)
        elif counter == 1 :
        	country_name = cell.value.strip()
        elif counter == 2:
        	production_type = cell.value.strip()
        elif cell.value is None or cell.value == "-" or cell.value == "--":
            prod_counter = prod_counter + 1
        else:
            cur_prod = int(cell.value)  
            year =  prod_counter / 12 + year_start
            month =   prod_counter % 12 + 1 
            if cur_prod > 0 :
                prod_obj = production.Production(country_id,production_type,cur_prod,year, month )
                prod_obj.save_to_db()
            prod_counter = prod_counter + 1
        counter = counter + 1
    print row_counter, 'rows processed'
    row_counter = row_counter + 1

        	